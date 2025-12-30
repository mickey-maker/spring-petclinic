import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

INPUT_FILE = "astf-python/combined_astf.json"
RAW_DATA_DIR = "astf-python/data"
OUTPUT_DIR = "astf-python/output"

SUPPRESS_LIST_CSV = "astf-python/suppress_list.csv"


# ===============================
# HELPERS
# ===============================
def _safe_upper(x) -> str:
    return str(x or "").upper().strip()


def _safe_strip(x) -> str:
    return str(x or "").strip()


def _line_to_int_or_none(x):
    if x in (None, "", "None"):
        return None
    try:
        return int(str(x).strip())
    except Exception:
        return None


def build_location(file_val: str, line_val) -> str:
    f = _safe_strip(file_val)
    ln = _line_to_int_or_none(line_val)
    if f and ln is not None:
        return f"{f}:{ln}"
    return f


# ===============================
# LOAD & NORMALISE ASTF DATA
# ===============================
def load_combined(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        print("[ERROR] combined_astf.json not found:", path)
        return pd.DataFrame()

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    if not raw:
        return pd.DataFrame()

    rows = []
    for item in raw:
        tool = _safe_upper(item.get("tool", ""))
        typ = _safe_upper(item.get("type", ""))
        sev = _safe_upper(item.get("severity", ""))
        msg = item.get("message", "")
        file_val = _safe_strip(item.get("file", ""))
        rule = _safe_strip(item.get("rule", ""))
        line_val = item.get("line", None)

        # Prefer location from combiner if present; otherwise derive it
        loc = _safe_strip(item.get("location", "")) or build_location(file_val, line_val)

        rows.append({
            "tool": tool,
            "type": typ,
            "severity": sev,
            "message": msg,
            "file": file_val,
            "rule": rule,
            "line": _line_to_int_or_none(line_val),
            "location": loc
        })

    print("[ASTF] ✅ ASTF alerts loaded:", len(rows))
    return pd.DataFrame(rows)


# ===============================
# DEDUPLICATION
# ===============================
def deduplicate_alerts(df: pd.DataFrame) -> pd.DataFrame:
    before = len(df)
    subset_cols = ["tool", "type", "rule", "location"]
    df = df.drop_duplicates(subset=subset_cols)
    after = len(df)
    print(f"[ASTF] ✅ Deduplication: {before} → {after}")
    return df


# ===============================
# PRIORITY SCORING
# ===============================
def apply_priority_scoring(df: pd.DataFrame) -> pd.DataFrame:
    # Severity weights
    SEV_WEIGHT = {"CRITICAL": 5, "HIGH": 4, "MEDIUM": 3, "LOW": 2, "INFO": 1}
    # Type weights (your ASTF goal is triage accuracy; vulnerability matters most)
    TYPE_WEIGHT = {"VULNERABILITY": 5, "BUG": 3, "CODE_SMELL": 1}

    df["sev_score"] = df["severity"].map(SEV_WEIGHT).fillna(1).astype(int)
    df["type_score"] = df["type"].map(TYPE_WEIGHT).fillna(1).astype(int)

    # Final Score formula (simple, explainable in Chapter 4/5)
    # final_score = severity_weight × type_weight
    df["final_score"] = (df["sev_score"] * df["type_score"]).astype(int)

    def assign(score: int) -> str:
        if score >= 20:
            return "P1"   # highest
        if score >= 12:
            return "P2"
        return "P3"

    df["priority"] = df["final_score"].apply(assign)
    print("[ASTF] ✅ Priority scoring completed")
    return df


# ===============================
# AUTO-GENERATE suppress_list.csv
# ===============================
def auto_generate_suppress_list(df: pd.DataFrame, output_path: str = SUPPRESS_LIST_CSV) -> pd.DataFrame:
    """
    Conservative suppression rules (noise proxy) to estimate FPR:
    - INFO severity => suppress
    - SAST CODE_SMELL => suppress
    - Priority P3 + LOW severity => suppress
    """
    suppressed_rows = []

    for _, row in df.iterrows():
        tool = row.get("tool", "")
        sev = row.get("severity", "")
        typ = row.get("type", "")
        pri = row.get("priority", "")
        rule = row.get("rule", "")
        file_ = row.get("file", "")
        line_ = row.get("line", None)
        location = row.get("location", build_location(file_, line_))

        reason = None
        if sev == "INFO":
            reason = "Informational severity (non-actionable)"
        elif pri == "P3" and sev in {"LOW", "INFO"}:
            reason = "Low priority noise proxy (P3 + LOW/INFO)"

        if reason:
            suppressed_rows.append({
                "tool": tool,
                "rule": rule,
                "file": file_,
                "line": line_,
                "location": location,   # ✅ exact location used for matching
                "reason": reason
            })

    sup_df = pd.DataFrame(suppressed_rows)
    if sup_df.empty:
        sup_df = pd.DataFrame(columns=["tool", "rule", "file", "line", "location", "reason"])

    # Stable matching keys: tool + rule + location
    sup_df["tool"] = sup_df["tool"].astype(str).str.upper().str.strip()
    sup_df["rule"] = sup_df["rule"].astype(str).str.strip()
    sup_df["location"] = sup_df["location"].astype(str).str.strip()

    sup_df = sup_df.drop_duplicates(subset=["tool", "rule", "location"])

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    sup_df.to_csv(output_path, index=False)
    print("[ASTF] ✅ suppress_list.csv auto-generated:", output_path)
    return sup_df


# ===============================
# APPLY SUPPRESSION
# ===============================
def apply_suppression(df: pd.DataFrame, suppress_csv: str = SUPPRESS_LIST_CSV) -> pd.DataFrame:
    df["suppressed"] = False
    df["suppress_reason"] = ""

    if not os.path.exists(suppress_csv):
        print("[ASTF] ⚠️ suppress_list.csv not found → suppression skipped.")
        print("[ASTF] ⚠️ expected at:", os.path.abspath(suppress_csv))
        return df

    sup = pd.read_csv(suppress_csv)

    required = ["tool", "rule", "location", "reason"]
    for col in required:
        if col not in sup.columns:
            raise ValueError(f"[ERROR] suppress_list.csv missing required column: {col}")

    # Normalize keys
    df["tool"] = df["tool"].astype(str).str.upper().str.strip()
    df["rule"] = df["rule"].astype(str).str.strip()
    df["location"] = df["location"].astype(str).str.strip()

    sup["tool"] = sup["tool"].astype(str).str.upper().str.strip()
    sup["rule"] = sup["rule"].astype(str).str.strip()
    sup["location"] = sup["location"].astype(str).str.strip()
    sup["reason"] = sup["reason"].astype(str).str.strip()

    df["__key"] = df["tool"] + "|" + df["rule"] + "|" + df["location"]
    sup["__key"] = sup["tool"] + "|" + sup["rule"] + "|" + sup["location"]

    sup_map = dict(zip(sup["__key"], sup["reason"]))

    df["suppressed"] = df["__key"].isin(sup_map.keys())
    df["suppress_reason"] = df["__key"].map(sup_map).fillna("")
    df.drop(columns=["__key"], inplace=True)

    print("[ASTF] ✅ Suppression applied:", int(df["suppressed"].sum()))
    return df


# ===============================
# TRIAGE METRICS (Expanded)
# ===============================
def generate_metrics_df(df: pd.DataFrame, raw_before_dedup: int = None) -> pd.DataFrame:
    total_after_dedup = len(df)
    suppressed = int(df["suppressed"].sum()) if "suppressed" in df.columns else 0
    actionable = total_after_dedup - suppressed

    if raw_before_dedup is None:
        raw_before_dedup = total_after_dedup

    dedup_removed = raw_before_dedup - total_after_dedup
    dedup_rate = (dedup_removed / raw_before_dedup * 100) if raw_before_dedup else 0.0

    suppression_rate = (suppressed / total_after_dedup * 100) if total_after_dedup else 0.0
    estimated_fpr = suppression_rate  # proxy definition (explain in thesis)

    # Distributions
    by_tool = df["tool"].value_counts().to_dict() if "tool" in df.columns else {}
    by_sev = df["severity"].value_counts().to_dict() if "severity" in df.columns else {}
    by_pri = df["priority"].value_counts().to_dict() if "priority" in df.columns else {}

    # Score stats
    score_mean = float(df["final_score"].mean()) if "final_score" in df.columns else 0.0
    score_median = float(df["final_score"].median()) if "final_score" in df.columns else 0.0
    score_max = int(df["final_score"].max()) if "final_score" in df.columns else 0

    rows = [
        ("Total Alerts (Raw, Before Dedup)", raw_before_dedup),
        ("Alerts After Deduplication", total_after_dedup),
        ("Deduplicated Alerts Removed", dedup_removed),
        ("Deduplication Rate (%)", round(dedup_rate, 2)),
        ("Suppressed Alerts", suppressed),
        ("Actionable Alerts", actionable),
        ("Suppression Rate (%)", round(suppression_rate, 2)),
        ("Estimated False Positive Rate (FPR) (%)", round(estimated_fpr, 2)),
        ("Mean Final Score", round(score_mean, 2)),
        ("Median Final Score", round(score_median, 2)),
        ("Max Final Score", score_max),
    ]

    # Priority breakdown
    for k in ["P1", "P2", "P3"]:
        rows.append((f"Priority Count ({k})", int(by_pri.get(k, 0))))

    # Tool breakdown
    for tool, cnt in by_tool.items():
        rows.append((f"Tool Count ({tool})", int(cnt)))

    # Severity breakdown
    for sev, cnt in by_sev.items():
        rows.append((f"Severity Count ({sev})", int(cnt)))

    return pd.DataFrame(rows, columns=["Metric", "Value"])


# ===============================
# RAW LISTS (for appendix / evidence)
# ===============================
def _pick_columns(df: pd.DataFrame, wanted: list) -> pd.DataFrame:
    existing = [c for c in wanted if c in df.columns]
    if not existing:
        return df.head(0)
    return df[existing].copy()


def load_sast_raw_main() -> pd.DataFrame:
    path = os.path.join(RAW_DATA_DIR, "sast.json")
    if not os.path.exists(path):
        print("[WARN] sast.json not found")
        return pd.DataFrame([{"error": "sast.json missing"}])

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    issues = raw.get("issues", []) if isinstance(raw, dict) else raw
    df = pd.json_normalize(issues)

    wanted = ["key", "rule", "severity", "type", "status", "component", "project", "line", "message", "creationDate", "updateDate"]
    return _pick_columns(df, wanted)


def load_dast_raw_main() -> pd.DataFrame:
    path = os.path.join(RAW_DATA_DIR, "dast.json")
    if not os.path.exists(path):
        print("[WARN] dast.json not found")
        return pd.DataFrame([{"error": "dast.json missing"}])

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    alerts = []
    if isinstance(raw, dict) and "site" in raw:
        for site in raw.get("site", []):
            alerts.extend(site.get("alerts", []))
    else:
        alerts = raw

    df = pd.json_normalize(alerts)
    wanted = ["pluginid", "alertRef", "alert", "riskcode", "confidence", "riskdesc", "desc", "solution", "reference", "cweid", "wascid", "sourceid"]
    return _pick_columns(df, wanted)


def load_sca_raw_main() -> pd.DataFrame:
    path = os.path.join(RAW_DATA_DIR, "sca.json")
    if not os.path.exists(path):
        print("[WARN] sca.json not found")
        return pd.DataFrame([{"error": "sca.json missing"}])

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    vulns = raw.get("vulnerabilities", []) if isinstance(raw, dict) else raw
    df = pd.json_normalize(vulns)
    wanted = ["id", "title", "severity", "cvssScore", "packageName", "moduleName", "language", "fixedIn", "patches"]
    return _pick_columns(df, wanted)


# ===============================
# SUMMARY DASHBOARD
# ===============================
def create_summary_dashboard(df: pd.DataFrame, excel_path: str):
    priority_summary = df["priority"].value_counts().reset_index()
    priority_summary.columns = ["PRIORITY", "COUNT"]

    tool_summary = df["tool"].value_counts().reset_index()
    tool_summary.columns = ["TOOL", "COUNT"]

    type_summary = df["type"].value_counts().reset_index()
    type_summary.columns = ["TYPE", "COUNT"]

    dashboard_rows = []
    dashboard_rows.append(["PRIORITY", "COUNT"])
    dashboard_rows.extend(priority_summary.values.tolist())
    dashboard_rows.append(["", ""])

    dashboard_rows.append(["TOOL", "COUNT"])
    dashboard_rows.extend(tool_summary.values.tolist())
    dashboard_rows.append(["", ""])

    dashboard_rows.append(["TYPE", "COUNT"])
    dashboard_rows.extend(type_summary.values.tolist())

    dashboard_df = pd.DataFrame(dashboard_rows, columns=["CATEGORY", "VALUE"])

    wb = load_workbook(excel_path)
    if "SUMMARY_DASHBOARD" in wb.sheetnames:
        del wb["SUMMARY_DASHBOARD"]
    wb.save(excel_path)

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a") as writer:
        dashboard_df.to_excel(writer, sheet_name="SUMMARY_DASHBOARD", index=False)

    print("[ASTF] ✅ SUMMARY_DASHBOARD updated")


# ===============================
# FORMATTING
# ===============================
def apply_global_formatting(excel_path: str):
    wb = load_workbook(excel_path)

    grey_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = grey_fill
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

    wb.save(excel_path)
    print("[ASTF] ✅ Formatting applied")


# ===============================
# SAVE EXCEL
# ===============================
def save_excel(astf_df: pd.DataFrame,
               metrics_df: pd.DataFrame,
               suppress_df: pd.DataFrame,
               sast_df: pd.DataFrame,
               dast_df: pd.DataFrame,
               sca_df: pd.DataFrame) -> str:

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, "astf_master_final.xlsx")

    astf_df = astf_df[astf_df["suppressed"] == False].copy()
    priority_rank = {"P1": 1, "P2": 2, "P3": 3}
    astf_df["priority_rank"] = astf_df["priority"].map(priority_rank).fillna(99)


    astf_df = astf_df.sort_values(
        by=["priority_rank", "final_score"],
        ascending=[True, False]
    ).drop(columns=["priority_rank"])

    astf_df = astf_df.reset_index(drop=True)


    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        astf_df.to_excel(writer, sheet_name="ASTF_FINAL", index=False)
        metrics_df.to_excel(writer, sheet_name="TRIAGE_METRICS", index=False)
        suppress_df.to_excel(writer, sheet_name="SUPPRESS_LIST", index=False)

        sast_df.to_excel(writer, sheet_name="SAST_RAW_LIST", index=False)
        dast_df.to_excel(writer, sheet_name="DAST_RAW_LIST", index=False)
        sca_df.to_excel(writer, sheet_name="SCA_RAW_LIST", index=False)

    print("[ASTF] ✅ Excel created:", output_path)
    print("[ASTF] ✅ Excel absolute path:", os.path.abspath(output_path))
    return output_path


# ===============================
# MAIN
# ===============================
def main():
    print("[ASTF] Working directory:", os.getcwd())
    print("[ASTF] combined_astf.json:", os.path.abspath(INPUT_FILE))
    print("[ASTF] suppress_list.csv:", os.path.abspath(SUPPRESS_LIST_CSV))

    df_raw = load_combined(INPUT_FILE)
    if df_raw.empty:
        print("[ASTF] ❌ No alerts detected. STOPPING.")
        return

    raw_before_dedup = len(df_raw)

    df = deduplicate_alerts(df_raw)
    df = apply_priority_scoring(df)

    # Generate suppress_list.csv automatically + keep a DF copy for Excel sheet
    suppress_df = auto_generate_suppress_list(df, SUPPRESS_LIST_CSV)

    # Apply suppression flags + reasons
    df = apply_suppression(df, SUPPRESS_LIST_CSV)

    metrics_df = generate_metrics_df(df, raw_before_dedup=raw_before_dedup)

    # Raw evidence sheets
    sast_df = load_sast_raw_main()
    dast_df = load_dast_raw_main()
    sca_df = load_sca_raw_main()

    excel_path = save_excel(df, metrics_df, suppress_df, sast_df, dast_df, sca_df)
    create_summary_dashboard(df, excel_path)
    apply_global_formatting(excel_path)

    print("\n✅ ASTF PIPELINE SUCCESSFUL")
    print("✅ Final ASTF Alerts (after dedup + scoring + suppression):", len(df))


if __name__ == "__main__":
    main()
